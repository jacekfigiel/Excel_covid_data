import openpyxl, pprint

print("Opening data...")
wb = openpyxl.load_workbook("COVID-19-geographic-disbtribution-worldwide-2020-12-14.xlsx")
sheet = wb.sheetnames
sheet1 = wb["COVID-19-geographic-disbtributi"]
country_data = {}

print("Reading data...")
for row in range(2, sheet1.max_row + 1):
    country = sheet1["G" + str(row)].value
    cases = sheet1["E" + str(row)].value
    deaths = sheet1["F" + str(row)].value

    country_data.setdefault(country, {"cases": 0,
                                      "deaths": 0})
    country_data[country]["cases"] += int(cases)
    country_data[country]["deaths"] += int(deaths)

print("Saving data...")
result_file = open("corona2021.py", "w")
result_file.write("all_data = " + pprint.pformat(country_data))
result_file.close()
print("Done!!!!")